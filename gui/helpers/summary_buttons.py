import os, datetime
from tkinter import filedialog, messagebox
from tkinter import ttk
from openpyxl import Workbook
from openpyxl.styles import Font

def build_button_bar(parent, controller):
    """
    Creates Close, Download CSV, and Process More Files buttons.
    """
    # Close
    close_btn = ttk.Button(
        parent,
        text="Close",
        style="Accent.TButton",
        command=controller.destroy
    )
    close_btn.pack(side="left")

    # Download CSV
    def on_download():
        mapping = getattr(controller, "final_mapping", {})
        # grab our timestamp dicts
        rename_ts = getattr(controller, "rename_timestamps", {})
        sheet_up   = getattr(controller, "sheet_updated", {})
        sheet_down = getattr(controller, "sheet_not_updated", {})

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Save summary report"
        )
        if not path:
            return

        wb = Workbook()
        # --- Sheet 1: Renamed Files ---
        ws1 = wb.active
        ws1.title = "Renamed Files"
        hdr = ["Old Filename", "New Filename", "Timestamp"]
        for col, name in enumerate(hdr, start=1):
            cell = ws1.cell(row=1, column=col, value=name)
            cell.font = Font(bold=True)
        for i, (old, new) in enumerate(mapping.items(), start=2):
            ts = rename_ts.get(old, datetime.datetime.now())
            ws1.cell(row=i, column=1, value=os.path.basename(old))
            ws1.cell(row=i, column=2, value=new)
            ws1.cell(row=i, column=3, value=ts.strftime("%Y-%m-%d %H:%M:%S"))

        # --- Sheet 2: Spreadsheet Changes ---
        ws2 = wb.create_sheet("Spreadsheet Changes")
        row = 1
        # Updated
        ws2.cell(row=row, column=1, value="Files UPDATED on Spreadsheet").font = Font(bold=True)
        row += 1
        for col, name in enumerate(hdr, start=1):
            ws2.cell(row=row, column=col, value=name).font = Font(bold=True)
        row += 1
        if sheet_up:
            for old in sheet_up:
                ts = sheet_up[old]
                ws2.cell(row=row, column=1, value=os.path.basename(old))
                ws2.cell(row=row, column=2, value=mapping.get(old, ""))
                ws2.cell(row=row, column=3, value=ts.strftime("%Y-%m-%d %H:%M:%S"))
                row += 1
        else:
            ws2.cell(row=row, column=1, value="No spreadsheet changes").font = Font(bold=True)
            row += 1

        # blank line
        row += 1
        # Not Updated
        ws2.cell(row=row, column=1, value="Files NOT UPDATED on Spreadsheet").font = Font(bold=True)
        row += 1
        for col, name in enumerate(hdr, start=1):
            ws2.cell(row=row, column=col, value=name).font = Font(bold=True)
        row += 1
        if sheet_down:
            for old in sheet_down:
                ts = sheet_down[old]
                ws2.cell(row=row, column=1, value=os.path.basename(old))
                ws2.cell(row=row, column=2, value=mapping.get(old, ""))
                ws2.cell(row=row, column=3, value=ts.strftime("%Y-%m-%d %H:%M:%S"))
                row += 1
        else:
            ws2.cell(row=row, column=1, value="All files were applied to the spreadsheet").font = Font(bold=True)

        # Save and notify
        try:
            wb.save(path)
            messagebox.showinfo("Saved", f"Report written to:\n{path}")
        except Exception as e:
            messagebox.showerror("Save Error", str(e))

    download_btn = ttk.Button(
        parent,
        text="Download CSV",
        style="Accent.TButton",
        command=on_download
    )
    download_btn.pack(side="left", padx=20)

    # Process More Files
    def on_more():
        # Reset per‑run state on the controller
        for attr in ("parse_index", "parse_mapping", "final_mapping", "ocr_results", "tif_list"):
            if hasattr(controller, attr):
                setattr(controller, attr, {} if "mapping" in attr else [])
        # Go back to folder page
        controller.show("folder_page")

    more_btn = ttk.Button(
        parent,
        text="Process More Files",
        style="Accent.TButton",
        command=on_more
    )
    more_btn.pack(side="right")
