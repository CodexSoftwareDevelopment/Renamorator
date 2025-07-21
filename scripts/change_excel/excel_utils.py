import os
import re
from typing import Dict, List, Tuple
from openpyxl import load_workbook

def update_spreadsheet(
    mapping: Dict[str, str],
    spreadsheet_path: str = 'data/Archived Protocol Status.xlsx',
    sheet_name: str = 'Document Scanning Tracker',
    doc_col_header: str = 'Document Name',
    number_col_header: str = 'Document Number',
    header_search_depth: int = 10,
    fuzzy_threshold: float = 0.9
) -> Tuple[int, List[str]]:
    """
    Updates the spreadsheet's Document Name column for each file in `mapping`
    where:
      1) The VP code(s) extracted from the new filename match the spreadsheet.
      2) The old filename stem matches fuzzily (>= fuzzy_threshold) the existing
         text in the Document Name column.

    If there are **multiple** VP codes in the new filename, it will write **all**
    of them (comma‑separated) into the Document Number column.

    Returns:
      - updates: the number of rows actually changed
      - unmatched: list of old filename stems that had no acceptable match
    """
    abs_path = os.path.abspath(spreadsheet_path)
    if not os.path.isfile(abs_path):
        raise FileNotFoundError(f"Spreadsheet not found: {abs_path}")

    # Load (and retry if locked)
    while True:
        try:
            wb = load_workbook(abs_path)
            break
        except PermissionError:
            input(f"Close '{abs_path}' if open, then press Enter to retry...")

    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    # Locate headers
    doc_col_idx = number_col_idx = None
    header_row = None
    for r in range(1, header_search_depth + 1):
        for c, cell in enumerate(ws[r], start=1):
            val = (cell.value or "").strip()
            if val == doc_col_header:
                doc_col_idx = c
                header_row = r
            elif val == number_col_header:
                number_col_idx = c
        if doc_col_idx and number_col_idx:
            break
    if not doc_col_idx or not number_col_idx:
        wb.close()
        raise ValueError(f"Headers '{doc_col_header}' and/or '{number_col_header}' not found in the first {header_search_depth} rows.")

    # Build a map: old_stem -> (new_stem, [vp_codes...])
    vp_re = re.compile(r"(C?VP\d{3,5}\.\d{3})", flags=re.IGNORECASE)
    stem_map: Dict[str, Tuple[str, List[str]]] = {}
    for old_path, new_path in mapping.items():
        old_stem = os.path.splitext(os.path.basename(old_path))[0]
        new_stem = os.path.splitext(os.path.basename(new_path))[0]
        vp_codes = vp_re.findall(new_stem)
        stem_map[old_stem] = (new_stem, vp_codes)

    updates = 0
    unmatched: List[str] = []
    last_row = ws.max_row

    # For each file, scan all rows until we find a match
    for old_stem, (new_stem, vp_codes) in stem_map.items():
        found = False
        for row in range(header_row + 1, last_row + 1):
            num_cell  = ws.cell(row, number_col_idx)
            name_cell = ws.cell(row, doc_col_idx)
            num_val   = (num_cell.value or "").strip()
            name_val  = (name_cell.value or "").strip()

            # We require that at least one of the VP codes matches num_val (or that
            # existing spreadsheet has a subset), but once we commit, we'll write all.
            # First do fuzzy name‑stem match:
            tokens      = re.findall(r"\w+", old_stem.lower())
            name_tokens = re.findall(r"\w+", name_val.lower())
            if not tokens:
                continue
            matches = sum(1 for t in tokens if t in name_tokens)
            ratio   = matches / len(tokens)
            if ratio < fuzzy_threshold:
                continue

            # OK name matched fuzzily—now verify that at least one vp_code is already there
            # or allow empty to fill in.  If multiple codes, we'll overwrite with all.
            if num_val and not any(code == num_val for code in vp_codes):
                # existing single-code mismatch; but we can still update if it's empty
                # or we choose to overwrite when codes differ.  For safety, only require
                # that num_val is empty or matches one code.
                continue

            # All conditions met: update!
            # 1) Write new document name
            name_cell.value = new_stem
            # 2) Write all vp_codes as comma-sep into Document Number
            num_cell.value = ", ".join(vp_codes)
            updates += 1
            found = True
            break

        if not found:
            unmatched.append(old_stem)

    if updates:
        wb.save(abs_path)
    wb.close()
    return updates, unmatched
